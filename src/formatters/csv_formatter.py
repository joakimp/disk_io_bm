"""CSV output formatter for benchmark results"""

import csv
from pathlib import Path
from typing import List


class CsvFormatter:
    """CSV output formatter"""

    def __init__(self, output_path: str):
        self.output_path = Path(output_path)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

    def format(self, results: List[dict]) -> None:
        """Format results as CSV and save to file"""
        if not results:
            return

        with open(self.output_path, "w", newline="") as f:
            writer = csv.writer(f)

            writer.writerow(
                [
                    "Test Type",
                    "Block Size",
                    "Read IOPS",
                    "Write IOPS",
                    "Read MB/s",
                    "Write MB/s",
                    "Read Lat (us)",
                    "Write Lat (us)",
                    "CPU",
                    "I/O Time (s)",
                    "Wall Time (s)",
                    "Status",
                ]
            )

            for result in results:
                # Support both old (runtime_sec) and new (io_time_sec) field names
                io_time = result.get("io_time_sec") or result.get("runtime_sec") or 0
                wall_time = result.get("wall_time_sec") or 0

                writer.writerow(
                    [
                        result.get("test_type", "N/A"),
                        result.get("block_size", "N/A"),
                        (result.get("read_iops") or 0),
                        (result.get("write_iops") or 0),
                        f"{(result.get('read_bw') or 0) / 1024 / 1024:.2f}",
                        f"{(result.get('write_bw') or 0) / 1024 / 1024:.2f}",
                        f"{(result.get('read_latency_us') or 0):.2f}",
                        f"{(result.get('write_latency_us') or 0):.2f}",
                        result.get("cpu", "N/A"),
                        f"{io_time:.2f}",
                        f"{wall_time:.2f}",
                        result.get("status", "N/A"),
                    ]
                )

        print(f"Results saved to {self.output_path}")


class ExcelFormatter:
    """Excel output formatter with multiple sheets"""

    def __init__(self, output_path: str):
        self.output_path = Path(output_path)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

    def format(self, results: List[dict]) -> None:
        """Format results as Excel with multiple sheets organized by metrics"""
        import pandas as pd

        if not results:
            print("No results to export")
            return

        df = pd.DataFrame(results)

        summary_df = df.groupby(["test_type", "block_size"]).agg(
            {
                "read_iops": ["mean", "min", "max"],
                "write_iops": ["mean", "min", "max"],
                "read_bw": ["mean", "min", "max"],
                "write_bw": ["mean", "min", "max"],
            }
        )
        summary_df.columns = ["_".join(col).strip() for col in summary_df.columns.values]
        summary_df = summary_df.reset_index()
        summary_df["Read MB/s"] = summary_df["read_bw_mean"] / 1024 / 1024
        summary_df["Write MB/s"] = summary_df["write_bw_mean"] / 1024 / 1024

        iops_df = df.pivot(
            index="block_size", columns="test_type", values=["read_iops", "write_iops"]
        )

        bw_df = df.copy()
        bw_df["Read MB/s"] = bw_df["read_bw"] / 1024 / 1024
        bw_df["Write MB/s"] = bw_df["write_bw"] / 1024 / 1024
        bw_df = bw_df.pivot(
            index="block_size", columns="test_type", values=["Read MB/s", "Write MB/s"]
        )

        lat_df = df.pivot(
            index="block_size",
            columns="test_type",
            values=["read_latency_us", "write_latency_us"],
        )

        # Support both old (runtime_sec) and new (io_time_sec, wall_time_sec) field names
        raw_columns = [
            "test_type",
            "block_size",
            "read_iops",
            "write_iops",
            "read_bw",
            "write_bw",
            "read_latency_us",
            "write_latency_us",
            "cpu",
        ]
        # Add time columns based on what's available
        if "io_time_sec" in df.columns:
            raw_columns.append("io_time_sec")
        elif "runtime_sec" in df.columns:
            raw_columns.append("runtime_sec")
        if "wall_time_sec" in df.columns:
            raw_columns.append("wall_time_sec")
        raw_columns.append("status")

        raw_df = df[[col for col in raw_columns if col in df.columns]]

        with pd.ExcelWriter(
            self.output_path, engine="openpyxl", datetime_format="YYYY-MM-DD HH:MM:SS"
        ) as writer:
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            iops_df.to_excel(writer, sheet_name="IOPS")
            bw_df.to_excel(writer, sheet_name="Bandwidth")
            lat_df.to_excel(writer, sheet_name="Latency")
            raw_df.to_excel(writer, sheet_name="Raw", index=False)

        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font

            workbook = load_workbook(self.output_path)
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.font = Font(bold=True)
            workbook.save(self.output_path)
        except Exception:
            pass

        print(f"Results saved to {self.output_path}")
